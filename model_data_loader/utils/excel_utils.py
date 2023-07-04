import re
from string import ascii_uppercase as auc
from xml.etree import ElementTree
from io import StringIO

from loguru import logger
from openpyxl.workbook import Workbook
from pandas import DataFrame


def excel_to_int(excel_index: str) -> int:
    """
    Преобразует индекс колокни Excel (А, B, C, ..., Z, AA, AB, ...) к int индексу.
    Индексу А соответсвует индекс 0.
    """
    num = 0
    for char in excel_index:
        num = num * 26 + (ord(char.upper()) - ord('A')) + 1
    return num - 1


def int_to_excel_index(num: int) -> str:
    """
    Преобразует int индекс к индексу колокни Excel (А, B, C, ..., Z, AA, AB, ...).
    Индексу 0 соответсвует индекс А.
    """
    if num > len(auc) - 1:
        return str(auc[num // 26 - 1]) + auc[num % 26]
    else:
        return auc[num % 26]


def load_xlsx_row(wb, sheet_name, row) -> list[str]:
    return load_xlsx_row_with_skip_list(wb, sheet_name, row, [])


def load_xlsx_row_with_skip_list(wb, sheet_name, row, skip_list) -> list[str]:
    """
    Parameters:
        wb: открытая книга workbook openoyxl
        sheet_name: название листа в книге
        row: номер ряда (нумерация с 1)
        skip_list: список подстрок, если значение в ряду содержит хотя бы одну подстроку из списка, оно будет пропущено
                    (на его месте будет пустая строка)
    Returns:
        Лист строк, содержащий значения ряда, которые не содержат ни одной подстроки из skip_list.
    """
    sheet_ranges = wb[sheet_name]
    if sheet_ranges is None:
        raise ValueError('Target sheet with name "' + sheet_name + '" was not found')
    res = []
    for cell in sheet_ranges[row]:
        line = cell.value
        if any(ext in str(line or '') for ext in skip_list):
            continue
        res.append(str(line or '').strip())
    return res


def get_column_offset(workbook, sheet_name, check_line_index) -> int:
    """
    Возвращает количество первых пустых ячеек в строке.
    Parameters:
        workbook: открытая книга openpyxl
        sheet_name: имя листа
        check_line_index: индекс строки, которая гарантированно не будет полностью пустой
    """
    i = 0
    while workbook[sheet_name][check_line_index][i].value is None:
        i += 1
    return i


def get_first_num_row_index(workbook, sheet_name, check_column_index) -> int:
    """
    Возвращает индекс первой строки с числовым или формульным форматом (нумерация с 0)
    Parameters:
        workbook: открытая книга openpyxl
        sheet_name: имя листа
        check_column_index: индекс колонки, которая гарантированно не будет полностью пустой
    """
    i = 1
    cell = workbook[sheet_name][i][check_column_index]
    while cell.value is None or (cell.data_type != 'n' and cell.data_type != 'f'):
        i += 1
        cell = workbook[sheet_name][i][check_column_index]
    return i - 1


# TODO возможно решение уже есть в openpyxl, но по нему очень мало документации:
#  https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.external_link.external.html
# TODO do more testing
def get_connections(extracted_zip: dict[str, bytes]) -> dict[str, str]:
    """
    Достает SQL запросы, объявленные в xlsx в разделе "Запросы и подключения"
    Возвращает словарик, в котором
        ключ: имя, которое упоминается в закладке "Область использования" в свойствах подключения
        значение: SQL запрос
    Parameters:
        extracted_zip: словарь, ключи - имена файлов внутри zip'а, значения - бинарные файлы, см. :fun: `extract_zip`
    """
    res = {}

    connections_xml_str = extracted_zip['xl/connections.xml'].decode("utf-8")
    connections_tree = ElementTree.fromstring(connections_xml_str)
    connections_namespace = \
        dict([node for _, node in ElementTree.iterparse(StringIO(connections_xml_str), events=['start-ns'])])
    connections = connections_tree.findall('.//{' + connections_namespace[''] + '}connection')

    table_list = [filename for filename in extracted_zip.keys() if re.match(r'xl/tables/table\d+\.xml', filename)]

    for table_name in table_list:
        table_xml_str = extracted_zip[table_name].decode("utf-8")
        table_xml = ElementTree.fromstring(table_xml_str)
        for connection in connections:
            if connection.attrib['id'] == table_xml.attrib['id']:
                res[table_xml.attrib['name']] = \
                    connection.find('.//{' + connections_namespace[''] + '}dbPr').attrib['command']
    return res


def update_sheet(workbook: Workbook, sheet_name: str, data: DataFrame, column_names: list[str], formulas: list[str],
                 row_offset: int, column_offset: int) -> None:
    """
    Заносит данные из data pandas DataFrame'a в лист sheet_name книги workbook,
    создает новую строку в конце, в которую сохраняет формулы из formulas
    Parameters:
        workbook: открытая книга openpyxl, по окончании не закрывается
        sheet_name: лист, в который пишем
        data: pandas DataFrame, названия колонок такие же как и в column_names (кол-во м.б. и меньше)
        column_names: названия всех колонок в листе sheet_name
        formulas: формулы для сохранения в последней строке
        row_offset: начальный адрес строки, нумерация с 0
        column_offset: начальный адрес колонки, нумерация с 0
    """
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    date_column_index = -1
    for (column_name, column_data) in data.items():
        column_index = column_names.index(column_name) + column_offset
        for j, value in enumerate(column_data):
            if column_name == 'Date':
                # TODO дата печатается как число
                date_column_index = column_index
                # j+1 т.к. в екселе строки нумеруются с 1
                sheet.cell(row=j + 1 + row_offset, column=column_index).number_format = 'dd.mm.yyyy'
                sheet.cell(row=j + 1 + row_offset, column=column_index).value = value
                # insert_value = value.strftime('%d.%m.%Y')
            else:
                insert_value = value
                sheet.cell(row=j + 1 + row_offset, column=column_index).value = insert_value

    # check if last row is already filled with formulas (date cell value is -1)
    last_date_cell_value = sheet.cell(row=sheet.max_row, column=date_column_index).value
    logger.debug(f'value in {date_column_index}:{sheet.max_row} cell is: '
                 f'{last_date_cell_value} with type {type(last_date_cell_value)}')
    logger.info('checking if last row contains formulas')
    if last_date_cell_value != -1:
        logger.info('creating last row with formulas')
        last_row_index = sheet.max_row + 1
        sheet.insert_rows(last_row_index)
        for column_index, formula in enumerate(formulas):
            sheet.cell(row=last_row_index, column=column_index + column_offset).value = formula
        sheet.cell(row=last_row_index, column=date_column_index).value = -1
    else:
        logger.info('date was matched')
