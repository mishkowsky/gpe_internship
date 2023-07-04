import os
import re
import shutil
import sys
from io import StringIO

from openpyxl.reader.excel import load_workbook
from pandas import DataFrame
import xml.etree.ElementTree as ElementTree
from db_loader import execute_query_to_dataframe
from formula_parser import FormulaParser
from config import TEMP_PATH, INPUT_PATH, OUTPUT_PATH
from utils.time_utils import get_date_range
from query_generator import generate_query
from utils.excel_utils import load_xlsx_row, get_column_offset, get_first_num_row_index, get_connections, update_sheet
from utils.zip_utils import extract_zip, insert_zip
from datetime import datetime, timedelta
from loguru import logger


# https://foss.heptapod.net/openpyxl/openpyxl/-/issues/2019
# TODO когда issue закроется, то можно будет сохранять конечный
#  файл с помощью openpyxl, не теряя "Запросы и подключения",
#  т.е. без плясок с архивациями
def process_model(input_file_path, temp_file_path, output_file_path,
         model_name, sheet_name, begin_date: datetime, end_date: datetime):
    """
    Parameters:
        input_file_path: путь к входному xlsx файлу
        temp_file_path: путь для хранения временной копии xlsx файла, включая имя
        output_file_path: путь к выходному xlsx файлу, включая имя
        sheet_name: имя листа
        begin_date:
        end_date:
    """

    if end_date - begin_date < timedelta(0):
        raise Exception("end_date can't be before begin_date")

    # TODO убрать из копии лишние листы, чтобы сократить время открытия/закрытия книги
    #  пока не удалось, если убирать лишние листы, забивая содержимое файлов нулями,
    #  то openpyxl жалуется, что xlsx битый
    # создаем копию книги
    shutil.copy(input_file_path, temp_file_path)

    logger.info('loading workbook')
    workbook = load_workbook(temp_file_path)

    column_offset = get_column_offset(workbook, sheet_name, 10)
    row_offset = get_first_num_row_index(workbook, sheet_name, 10)

    logger.info('workbook loaded, parsing formulas')
    # читаем формулы из последней строки
    formulas = load_xlsx_row(workbook, sheet_name, workbook[sheet_name].max_row)
    logger.info(f'loaded formulas from {workbook[sheet_name].max_row}')

    # адрес строки с именами колонок row_offset-1 и +1 тк нумерация строк с 1
    column_names = load_xlsx_row(workbook, sheet_name, row_offset)

    # разархивируем копию in-memory
    temporary_extracted = extract_zip(temp_file_path)

    # парсим формулы екселя в python объекты
    formula_parser = FormulaParser(formulas[:], get_connections(temporary_extracted), model_name)

    logger.info('formulas parsed, generating query')
    # генерируем запрос, результат запроса помещаем в pandas DataFrame
    query = generate_query(
        list(formula_parser.data_sources.values()), formula_parser.sum_if_formulas, column_names, begin_date, end_date
    )
    logger.debug('generated query:\n' + query)

    logger.info('query generated, executing query')
    df_generated: DataFrame = execute_query_to_dataframe(query)

    # В копию книги в лист 'sheet_name' вносим изменения
    logger.info(f'query executed, updating sheet {sheet_name}')
    update_sheet(workbook, sheet_name, df_generated, column_names, formulas, row_offset, column_offset)

    logger.info(f'sheet {sheet_name} updated, saving')
    workbook.save(temp_file_path)

    # разархивируем in-memory только что сохранненый xlsx
    temporary_extracted = extract_zip(temp_file_path)

    # Копию листа из временной книги вставляем в исходную
    workbook_xml_str = temporary_extracted['xl/workbook.xml'].decode("utf-8")

    # взято с https://stackoverflow.com/a/42338368
    namespace = dict([node for _, node in ElementTree.iterparse(StringIO(workbook_xml_str), events=['start-ns'])])

    tree = ElementTree.fromstring(workbook_xml_str)
    sheets = tree.findall('.//{' + namespace[''] + '}sheet')

    target_xml = None
    target_xml_name = ''
    for sheet in sheets:
        if sheet.attrib['name'] == sheet_name:
            sheet_r_id = sheet.attrib['{' + namespace['r'] + '}id']
            matcher = re.search(r'\d+$', sheet_r_id)
            if not matcher:
                raise ValueError('cant find index')
            sheet_index = matcher.group()
            target_xml_name = 'xl/worksheets/sheet' + sheet_index + '.xml'
            target_xml = temporary_extracted[target_xml_name]
    if target_xml is None:
        raise ValueError('Target list "Daily" was not found')

    # разархивируем входной файл, чтобы внести изменения в него
    main_extracted = extract_zip(input_file_path)
    main_extracted[target_xml_name] = target_xml

    insert_zip(output_file_path, main_extracted)

    os.remove(temp_file_path)
    logger.info(f'finished processing {model_name}')


def main():

    logger.info('starting')

    if not os.path.exists('./' + TEMP_PATH):
        os.makedirs('./' + TEMP_PATH)

    for file_name in os.listdir(INPUT_PATH):
        if not re.match(r'.*\.xlsx$', file_name):
            logger.debug(f'skipping {file_name}')
            continue
        matcher = re.match(r'(.*)(?=\.xlsx)', file_name)
        logger.info(f'processing {file_name} file')
        if matcher:
            model_name = matcher.group(0)
        else:
            raise RuntimeError(f'empty file name: {file_name}')
        # TODO what's range to use?
        begin_date, end_date = get_date_range()
        logger.debug('date range: from ' + begin_date.strftime('%d.%m.%Y') + ' to ' + end_date.strftime('%d.%m.%Y'))
        process_model(
            input_file_path=INPUT_PATH + '/' + file_name,
            temp_file_path=TEMP_PATH + '/' + file_name,
            output_file_path=OUTPUT_PATH + '/' + file_name,
            model_name=model_name,
            sheet_name='Daily',
            begin_date=begin_date, end_date=end_date
        )
    logger.info('finished processing files')


if __name__ == '__main__':
    logger.remove()
    logger.add(sys.stderr, level="INFO")
    main()
